"""
dashboard/routes.py — All dashboard routes (v4 complete)
"""
from flask import Blueprint, render_template, jsonify, request, session, redirect, url_for, send_file
import io, json
from core.database import (get_recent_trades, get_recent_webhooks, get_closed_positions,
                            get_closed_summary, log_closed_position)
from core.config import Config
from core.telegram import send_telegram, alert_kill_switch
from core.excel_export import export_trades_excel
from brokers.alpaca_adapter import AlpacaAdapter
from core.logger import get_logger

logger = get_logger(__name__)
dashboard_bp = Blueprint("dashboard", __name__)
alpaca = AlpacaAdapter()

def _auth():
    if not session.get("logged_in"):
        return jsonify({"error": "Unauthorized"}), 401
    return None

@dashboard_bp.route("/")
def index():
    if not session.get("logged_in"):
        return redirect(url_for("dashboard.login"))
    # Pre-load portfolio data server-side so Analytics tab never needs a fetch
    import json, requests as req
    from core.config import Config
    import os
    portfolio_json = "null"
    try:
        key = (Config.ALPACA_API_KEY
               or os.environ.get("APCA_API_KEY_ID","")
               or os.environ.get("ALPACA_KEY",""))
        sec = (Config.ALPACA_SECRET_KEY
               or os.environ.get("APCA_API_SECRET_KEY","")
               or os.environ.get("ALPACA_SECRET",""))
        hdrs = {"APCA-API-KEY-ID": key, "APCA-API-SECRET-KEY": sec}
        base = Config.ALPACA_BASE_URL
        acc_r = req.get(f"{base}/v2/account",   headers=hdrs, timeout=8)
        pos_r = req.get(f"{base}/v2/positions",  headers=hdrs, timeout=8)
        if acc_r.ok and pos_r.ok:
            acc = acc_r.json()
            raw = pos_r.json()
            equity = float(acc.get("portfolio_value") or acc.get("equity") or 0)
            cash   = float(acc.get("cash") or 0)
            SECTOR = {
                "AAPL":"Technology","MSFT":"Technology","NVDA":"Technology","AMD":"Technology",
                "MU":"Technology","INTC":"Technology","LRCX":"Technology","IBM":"Technology",
                "MRVL":"Technology","ORCL":"Technology","PLTR":"Technology","SMCI":"Technology",
                "QCOM":"Technology","AVGO":"Technology","TXN":"Technology","AMAT":"Technology",
                "CRM":"Technology","SNOW":"Technology","PANW":"Technology","CRWD":"Technology",
                "META":"Communication","GOOGL":"Communication","GOOG":"Communication",
                "NFLX":"Communication","DIS":"Communication","T":"Communication","VZ":"Communication",
                "AMZN":"Consumer Disc.","TSLA":"Consumer Disc.","HD":"Consumer Disc.",
                "NKE":"Consumer Disc.","SBUX":"Consumer Disc.","BKNG":"Consumer Disc.",
                "JPM":"Financials","BAC":"Financials","GS":"Financials","MS":"Financials",
                "V":"Financials","MA":"Financials","PYPL":"Financials","SQ":"Financials",
                "COIN":"Financials","HOOD":"Financials","MARA":"Financials","RIOT":"Financials",
                "JNJ":"Healthcare","UNH":"Healthcare","PFE":"Healthcare","MRNA":"Healthcare",
                "LLY":"Healthcare","ABT":"Healthcare","ISRG":"Healthcare","GILD":"Healthcare",
                "XOM":"Energy","CVX":"Energy","COP":"Energy","SLB":"Energy",
                "SPY":"ETF","QQQ":"ETF","IWM":"ETF","GLD":"ETF","TLT":"ETF",
                "ARKK":"ETF","TQQQ":"ETF","SQQQ":"ETF","DIA":"ETF",
                "BA":"Industrials","CAT":"Industrials","GE":"Industrials","LMT":"Industrials",
            }
            positions, sector_totals, total_mv, total_upl, total_cb = [], {}, 0.0, 0.0, 0.0
            for p in raw:
                sym  = p.get("symbol","")
                qty  = float(p.get("qty") or 0)
                mv   = float(p.get("market_value") or 0)
                cb   = float(p.get("cost_basis") or 0)
                upl  = float(p.get("unrealized_pl") or 0)
                uplpc= float(p.get("unrealized_plpc") or 0) * 100
                entry= float(p.get("avg_entry_price") or 0)
                cp   = float(p.get("current_price") or 0)
                sec  = SECTOR.get(sym.upper(), "Other")
                alloc= round(mv / equity * 100, 2) if equity else 0
                total_mv += mv; total_upl += upl; total_cb += cb
                positions.append({"symbol":sym,"qty":qty,"side":p.get("side","long"),
                    "avg_entry":round(entry,4),"current_price":round(cp,4),
                    "market_value":round(mv,2),"cost_basis":round(cb,2),
                    "unrealized_pl":round(upl,2),"unrealized_pct":round(uplpc,3),
                    "sector":sec,"allocation_pct":alloc})
                if sec not in sector_totals:
                    sector_totals[sec] = {"market_value":0,"unrealized_pl":0,"count":0,"symbols":[]}
                sector_totals[sec]["market_value"]  += mv
                sector_totals[sec]["unrealized_pl"] += upl
                sector_totals[sec]["count"]         += 1
                sector_totals[sec]["symbols"].append(sym)
            positions.sort(key=lambda x: x["unrealized_pl"], reverse=True)
            sec_bd = {s: {"market_value":round(d["market_value"],2),
                          "unrealized_pl":round(d["unrealized_pl"],2),
                          "allocation_pct":round(d["market_value"]/equity*100,2) if equity else 0,
                          "count":d["count"],"symbols":d["symbols"]}
                      for s,d in sorted(sector_totals.items(), key=lambda x:-x[1]["market_value"])}
            best  = max(positions, key=lambda x:x["unrealized_pct"]) if positions else {}
            worst = min(positions, key=lambda x:x["unrealized_pct"]) if positions else {}
            payload = {
                "account":{"equity":round(float(acc.get("equity",0)),2),
                            "cash":round(cash,2),
                            "buying_power":round(float(acc.get("buying_power",0)),2),
                            "portfolio_value":round(equity,2)},
                "positions": positions,
                "sector_breakdown": sec_bd,
                "summary":{"open_count":len(positions),
                           "total_market_value":round(total_mv,2),
                           "total_unrealized_pl":round(total_upl,2),
                           "total_cost_basis":round(total_cb,2),
                           "roi_pct":round(total_upl/total_cb*100,3) if total_cb else 0,
                           "invested_pct":round(total_mv/equity*100,2) if equity else 0,
                           "cash_pct":round(cash/equity*100,2) if equity else 0,
                           "best_symbol":best.get("symbol","—"),
                           "best_pct":best.get("unrealized_pct",0),
                           "best_pl":best.get("unrealized_pl",0),
                           "worst_symbol":worst.get("symbol","—"),
                           "worst_pct":worst.get("unrealized_pct",0),
                           "worst_pl":worst.get("unrealized_pl",0)},
            }
            portfolio_json = json.dumps(payload)
    except Exception as ex:
        portfolio_json = json.dumps({"error": str(ex), "positions":[], "summary":{}, "sector_breakdown":{}, "account":{}})
    return render_template("dashboard.html", portfolio_json=portfolio_json)

@dashboard_bp.route("/terminal")
def terminal():
    if not session.get("logged_in"):
        return redirect(url_for("dashboard.login"))
    return render_template("terminal.html")

@dashboard_bp.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == Config.DASHBOARD_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard.index"))
        error = "Wrong password"
    return render_template("login.html", error=error)

@dashboard_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("dashboard.login"))

@dashboard_bp.route("/api/account")
def api_account():
    e = _auth()
    if e: return e
    try:
        account   = alpaca.get_account()
        positions = alpaca.get_positions()

        # Calculate open positions totals
        total_market_value = 0.0
        total_unrealized_pl = 0.0
        for p in positions:
            try:
                total_market_value  += float(p.get("market_value") or 0)
                total_unrealized_pl += float(p.get("unrealized_pl") or 0)
            except:
                pass

        return jsonify({
            "account":            account,
            "positions":          positions,
            "total_market_value": round(total_market_value, 2),
            "total_unrealized_pl":round(total_unrealized_pl, 2),
            "open_count":         len(positions),
        })
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/api/trades")
def api_trades():
    e = _auth()
    if e: return e
    # Sync from Alpaca first so we always show latest
    try:
        from core.order_sync import sync_alpaca_orders
        sync_alpaca_orders(days=7)
    except Exception as ex:
        logger.warning(f"Order sync skipped: {ex}")
    return jsonify(get_recent_trades(200))

@dashboard_bp.route("/api/webhooks")
def api_webhooks():
    e = _auth(); 
    if e: return e
    return jsonify(get_recent_webhooks(20))

@dashboard_bp.route("/api/closed_positions")
def api_closed_positions():
    e = _auth()
    if e: return e
    # Sync latest orders to capture any new closed trades
    try:
        from core.order_sync import sync_alpaca_orders
        sync_alpaca_orders(days=30)
    except Exception as ex:
        logger.warning(f"Order sync skipped: {ex}")
    return jsonify({
        "positions": get_closed_positions(200),
        "summary":   get_closed_summary(),
    })

@dashboard_bp.route("/api/close_position", methods=["POST"])
def api_close_position():
    e = _auth(); 
    if e: return e
    data   = request.json or {}
    symbol = data.get("symbol","").upper().strip()
    qty    = data.get("qty", None)
    if not symbol:
        return jsonify({"success": False, "error": "Symbol required"}), 400
    try:
        pos          = alpaca.get_position(symbol)
        entry_price  = float(pos.get("avg_entry_price", 0)) if pos else None
        current_price= float(pos.get("current_price",  0)) if pos else None
        pos_qty      = float(pos.get("qty", 0))             if pos else 0
        side         = "long" if pos_qty > 0 else "short"

        if qty and float(qty) < abs(pos_qty):
            close_qty  = float(qty)
            order_side = "sell" if pos_qty > 0 else "buy"
            result     = alpaca.place_market_order(symbol, order_side, close_qty)
        else:
            result    = alpaca.close_position(symbol)
            close_qty = abs(pos_qty)

        if entry_price and current_price:
            log_closed_position(symbol, close_qty, entry_price, current_price, side,
                                alpaca_id=result.get("id") if result else None)
        pnl = ((current_price - entry_price) * close_qty if side == "long"
               else (entry_price - current_price) * close_qty) if entry_price and current_price else 0
        send_telegram(f"📤 <b>CLOSED {symbol}</b> Qty:{close_qty} P&L:<b>${pnl:+.2f}</b>")
        return jsonify({"success": True, "pnl": round(pnl,2)})
    except Exception as ex:
        return jsonify({"success": False, "error": str(ex)}), 500

@dashboard_bp.route("/api/close_all", methods=["POST"])
def api_close_all():
    e = _auth(); 
    if e: return e
    try:
        positions = alpaca.get_positions()
        for p in positions:
            sym  = p["symbol"]
            qty  = float(p["qty"])
            en   = float(p.get("avg_entry_price",0))
            cu   = float(p.get("current_price",0))
            side = "long" if qty > 0 else "short"
            log_closed_position(sym, abs(qty), en, cu, side)
        result = alpaca.close_all_positions()
        send_telegram("🚨 <b>ALL POSITIONS CLOSED</b> via dashboard")
        return jsonify({"success": True})
    except Exception as ex:
        return jsonify({"success": False, "error": str(ex)}), 500

@dashboard_bp.route("/api/kill_switch", methods=["POST"])
def api_kill_switch():
    e = _auth(); 
    if e: return e
    state = request.json.get("enabled", True)
    with open(".kill_switch","w") as f:
        f.write("1" if state else "0")
    Config.KILL_SWITCH = state
    alert_kill_switch(state)
    return jsonify({"success": True, "kill_switch": state})

@dashboard_bp.route("/api/export_excel")
def api_export_excel():
    e = _auth(); 
    if e: return e
    try:
        data = export_trades_excel()
        from datetime import datetime
        return send_file(io.BytesIO(data),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"trades_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/api/test_telegram", methods=["POST"])
def api_test_telegram():
    e = _auth(); 
    if e: return e
    ok = send_telegram("✅ <b>Telegram connected!</b>\nYour OptiTrade bot is online.")
    return jsonify({"success": ok, "message": "Sent!" if ok else "Failed — check TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID"})

# ── Analyzer Pro ──────────────────────────────────────────────────────────────
@dashboard_bp.route("/api/analyze", methods=["POST"])
def api_analyze():
    e = _auth(); 
    if e: return e
    data       = request.json or {}
    symbols    = [s.strip().upper() for s in data.get("symbols",[]) if s.strip()][:10]
    timeframes = data.get("timeframes", ["15m","1h","1D"])
    if not symbols:
        return jsonify({"error":"No symbols"}), 400
    try:
        from core.analyzer import analyze_multiple
        results = analyze_multiple(symbols, timeframes)
        from core.market_data import data_source_status
        return jsonify({"results": results, "timeframes": timeframes,
                        "data_source": data_source_status()})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

# ── Stock chart data ──────────────────────────────────────────────────────────
@dashboard_bp.route("/api/chart_data", methods=["POST"])
def api_chart_data():
    e = _auth()
    if e: return e
    data   = request.json or {}
    symbol = data.get("symbol","").upper()
    period = data.get("period","3mo")
    try:
        from core.market_data import get_bars, PERIOD_CONFIG
        from datetime import datetime

        bars = get_bars(symbol, period)
        if not bars:
            return jsonify({"error": "No data available"}), 404

        # ── Smart date label formatting based on period ─────────────────────
        # Intraday periods need time, not just date
        intraday_periods = ("5m","15m","1h","4h")
        long_periods     = ("3y","5y","10y","3Y","5Y","10Y")
        mid_periods      = ("1y","1W","1D")

        period_labels = {
            "5m":"5 Min","15m":"15 Min","1h":"1 Hour","4h":"4 Hour",
            "1mo":"1 Month","3mo":"3 Months","6mo":"6 Months",
            "1y":"1 Year","1D":"Daily","1W":"Weekly",
            "3y":"3 Years","5y":"5 Years","10y":"10 Years",
        }

        dates = []
        for b in bars:
            try:
                dt = datetime.fromisoformat(b["t"].replace("Z","+00:00"))
                if period in intraday_periods:
                    # Show date only at day boundaries, otherwise just time
                    dates.append(dt.strftime("%m/%d %H:%M"))
                elif period in long_periods:
                    dates.append(dt.strftime("%b %Y"))
                elif period in mid_periods:
                    dates.append(dt.strftime("%b %d '%y"))
                else:
                    # 1mo, 3mo, 6mo — daily, show month+day
                    dates.append(dt.strftime("%b %d"))
            except:
                dates.append("")

        closes = [b.get("c") for b in bars]

        def ema_series(data, span):
            if not data or len(data) < span: return []
            k = 2/(span+1)
            s = next((x for x in data if x), data[0])
            result = []
            for p in data:
                if p is None: result.append(None); continue
                s = p*k + s*(1-k)
                result.append(round(s, 2))
            return result

        c_valid = [c for c in closes if c is not None]
        ema9   = ema_series(c_valid, 9)
        ema21  = ema_series(c_valid, 21)
        ema63  = ema_series(c_valid, 63)
        ema200 = ema_series(c_valid, 200) if len(c_valid) >= 200 else []

        # Max tick limit based on period for clean X-axis
        max_ticks = 8
        if period in intraday_periods:
            max_ticks = 12
        elif period in ("1mo",):
            max_ticks = 8
        elif period in ("3mo","6mo"):
            max_ticks = 10
        elif period in long_periods:
            max_ticks = 8

        return jsonify({
            "symbol":     symbol,
            "period":     period,
            "period_label": period_labels.get(period, period),
            "dates":      dates,
            "open":       [b.get("o") for b in bars],
            "high":       [b.get("h") for b in bars],
            "low":        [b.get("l") for b in bars],
            "close":      closes,
            "volume":     [b.get("v") for b in bars],
            "ema9":       ema9, "ema21": ema21,
            "ema63":      ema63, "ema200": ema200,
            "source":     bars[0].get("source","demo") if bars else "demo",
            "bar_count":  len(bars),
            "max_ticks":  max_ticks,
            "is_intraday": period in intraday_periods,
            "date_range": f"{dates[0] if dates else ''} → {dates[-1] if dates else ''}",
        })
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

# ── Compare stocks ────────────────────────────────────────────────────────────
@dashboard_bp.route("/api/compare", methods=["POST"])
def api_compare():
    e = _auth()
    if e: return e
    data    = request.json or {}
    symbols = [s.strip().upper() for s in data.get("symbols",[]) if s.strip()][:8]
    period  = data.get("period","3mo")
    if not symbols:
        return jsonify({"error":"No symbols"}), 400
    try:
        from core.market_data import get_bars
        from datetime import datetime
        intraday_periods = ("5m","15m","1h","4h")
        long_p  = ("3y","5y","10y","3Y","5Y","10Y")
        mid_p   = ("1y","1W","1D")
        result  = {}
        for sym in symbols:
            bars = get_bars(sym, period)
            if not bars: continue
            dates, closes = [], []
            for b in bars:
                try:
                    dt = datetime.fromisoformat(b["t"].replace("Z","+00:00"))
                    if period in intraday_periods:
                        dates.append(dt.strftime("%m/%d %H:%M"))
                    elif period in long_p:
                        dates.append(dt.strftime("%b %Y"))
                    elif period in mid_p:
                        dates.append(dt.strftime("%b %d '%y"))
                    else:
                        dates.append(dt.strftime("%b %d"))
                except:
                    dates.append("")
                closes.append(b.get("c"))
            base = next((c for c in closes if c), None)
            if not base: continue
            norm = [round((c/base-1)*100, 2) if c else None for c in closes]
            result[sym] = {
                "dates":  dates, "norm": norm, "closes": closes,
                "source": bars[0].get("source","demo") if bars else "demo",
                "period": period, "bar_count": len(bars),
                "is_intraday": period in intraday_periods,
                "date_range": f"{dates[0] if dates else ''} → {dates[-1] if dates else ''}",
            }
        return jsonify(result)
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

# ── Research routes (free data) ───────────────────────────────────────────────
@dashboard_bp.route("/api/research/institutional")
def api_research_institutional():
    e = _auth(); 
    if e: return e
    try:
        from research.sec_filings import get_institutional_tracker, analyze_institutional_momentum
        data      = get_institutional_tracker()
        momentum  = analyze_institutional_momentum(data)
        return jsonify({"funds": data, "momentum_stocks": momentum})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/api/research/earnings")
def api_research_earnings():
    e = _auth(); 
    if e: return e
    try:
        from research.earnings import get_earnings_whiplash
        return jsonify(get_earnings_whiplash(max_stocks=50))
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/api/research/sectors")
def api_research_sectors():
    e = _auth(); 
    if e: return e
    try:
        from research.sector_rotation import get_sector_rotation
        return jsonify(get_sector_rotation())
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/api/research/insider", methods=["POST"])
def api_research_insider():
    e = _auth(); 
    if e: return e
    data    = request.json or {}
    symbols = data.get("symbols", None)
    try:
        from research.insider_flow import get_confluence_stocks
        return jsonify(get_confluence_stocks(symbols))
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

# ── Telegram bot commands ─────────────────────────────────────────────────────
@dashboard_bp.route("/telegram/webhook", methods=["POST"])
def telegram_webhook():
    try:
        data    = request.get_json(force=True, silent=True) or {}
        msg     = data.get("message",{})
        text    = msg.get("text","").strip()
        chat_id = str(msg.get("chat",{}).get("id",""))
        if chat_id != Config.TELEGRAM_CHAT_ID:
            return jsonify({"ok":True})
        cmd = text.lower().split()[0] if text else ""
        if cmd in ["/positions","/pos"]:
            positions = alpaca.get_positions()
            if not positions:
                send_telegram("📊 <b>No open positions</b>")
            else:
                lines = ["📊 <b>Open Positions</b>\n━━━━━━━━━━━━━━"]
                for p in positions:
                    pnl = float(p.get("unrealized_pl",0))
                    pct = float(p.get("unrealized_plpc",0))*100
                    lines.append(f"📌 <b>{p['symbol']}</b> Qty:{p['qty']}\n"
                                 f"   P&L: <b>${pnl:+.2f} ({pct:+.2f}%)</b>")
                send_telegram("\n".join(lines))
        elif cmd in ["/pnl","/summary"]:
            from core.database import get_closed_summary
            s  = get_closed_summary()
            wr = (s["winners"]/s["total_trades"]*100) if s.get("total_trades") else 0
            send_telegram(f"📋 <b>P&L Summary</b>\n━━━━━━━━━━━━━━\n"
                          f"Trades:{s.get('total_trades',0)} WinRate:<b>{wr:.1f}%</b>\n"
                          f"Total P&L:<b>${s.get('total_pnl',0):+.2f}</b>")
        elif cmd in ["/account","/balance"]:
            acc = alpaca.get_account()
            pnl = float(acc.get("equity",0)) - float(acc.get("last_equity",0))
            send_telegram(f"💼 <b>Account</b>\n"
                          f"Portfolio: <b>${float(acc.get('portfolio_value',0)):,.2f}</b>\n"
                          f"Today P&L: <b>${pnl:+.2f}</b>")
        elif cmd == "/closeall":
            alpaca.close_all_positions()
            send_telegram("🚨 <b>ALL POSITIONS CLOSED</b>")
        elif cmd == "/close" and len(text.split()) > 1:
            sym = text.split()[1].upper()
            try:
                alpaca.close_position(sym)
                send_telegram(f"✅ <b>{sym}</b> closed.")
            except Exception as ex:
                send_telegram(f"❌ Failed: {str(ex)[:100]}")
        else:
            send_telegram("🤖 Commands: /positions /pnl /account /close AAPL /closeall")
        return jsonify({"ok":True})
    except Exception as ex:
        logger.error(f"Telegram webhook: {ex}")
        return jsonify({"ok":True})

@dashboard_bp.route("/api/clock")
def api_clock():
    """Return current time in UTC, NY, JST + market session schedule."""
    try:
        from core.timezone_utils import now_all_zones, market_sessions_utc
        t    = now_all_zones()
        sess = market_sessions_utc()
        return jsonify({
            "times":    t,
            "sessions": sess["sessions"],   # flat dict of session ranges
            "tz_label": sess["tz_label"],
            "ny_offset":sess["ny_offset"],
        })
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@dashboard_bp.route("/health")
def health():
    return jsonify({"status":"ok","message":"Trading bot running"})

@dashboard_bp.route("/api/data_status")
def api_data_status():
    from core.market_data import data_source_status
    return jsonify(data_source_status())

# ── Market Sessions (Asia/London/New York) ────────────────────────────────────
@dashboard_bp.route("/api/sessions", methods=["POST"])
def api_sessions():
    e = _auth();
    if e: return e
    data    = request.json or {}
    symbol  = data.get("symbol","SPY").upper()
    try:
        from core.market_data import get_bars
        from datetime import datetime, timezone
        import math

        # Get intraday 1h bars (last 5 days)
        bars = get_bars(symbol, "1h")
        if not bars:
            return jsonify({"error": "No intraday data"}), 404

        # Session times UTC
        # Asia:    00:00 - 08:00 UTC
        # London:  08:00 - 13:00 UTC
        # NY:      13:30 - 20:00 UTC
        sessions = {"Asia": [], "London": [], "NewYork": []}

        for b in bars[-120:]:  # last 5 days of hourly
            try:
                dt   = datetime.fromisoformat(b["t"].replace("Z","+00:00"))
                hour = dt.hour
                day  = dt.strftime("%Y-%m-%d")
                bar_data = {
                    "time":  dt.strftime("%m/%d %H:%M"),
                    "day":   day,
                    "open":  b.get("o"), "high": b.get("h"),
                    "low":   b.get("l"), "close": b.get("c"),
                    "vol":   b.get("v",0),
                }
                if 0 <= hour < 8:
                    sessions["Asia"].append(bar_data)
                elif 8 <= hour < 13:
                    sessions["London"].append(bar_data)
                elif 13 <= hour < 21:
                    sessions["NewYork"].append(bar_data)
            except:
                continue

        def session_stats(bars_list):
            if not bars_list: return None
            # Group by day
            days = {}
            for b in bars_list:
                d = b["day"]
                if d not in days: days[d] = []
                days[d].append(b)

            day_stats = []
            for day, dbars in list(days.items())[-5:]:
                opens  = [x["open"]  for x in dbars if x["open"]]
                closes = [x["close"] for x in dbars if x["close"]]
                highs  = [x["high"]  for x in dbars if x["high"]]
                lows   = [x["low"]   for x in dbars if x["low"]]
                vols   = [x["vol"]   for x in dbars if x["vol"]]
                if not opens or not closes: continue
                o = opens[0]; c = closes[-1]
                h = max(highs) if highs else c
                l = min(lows)  if lows  else c
                chg = round((c-o)/o*100,2) if o else 0
                rng = round(h-l, 2)
                vol = sum(vols)
                day_stats.append({
                    "date": day, "open": round(o,2), "close": round(c,2),
                    "high": round(h,2), "low": round(l,2),
                    "change_pct": chg, "range": rng, "volume": vol,
                    "bias": "Accumulation" if chg > 0.3 else
                            "Distribution"  if chg < -0.3 else "Consolidation",
                    "bars": dbars,
                })
            if not day_stats: return None
            avg_chg = sum(d["change_pct"] for d in day_stats)/len(day_stats)
            return {
                "days": day_stats,
                "avg_change": round(avg_chg,2),
                "dominant_bias": "Accumulation" if avg_chg > 0.2 else
                                 "Distribution"  if avg_chg < -0.2 else "Consolidation",
            }

        result = {
            "symbol":   symbol,
            "source":   bars[0].get("source","demo") if bars else "demo",
            "generated":datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "Asia":     session_stats(sessions["Asia"]),
            "London":   session_stats(sessions["London"]),
            "NewYork":  session_stats(sessions["NewYork"]),
        }

        # Cross-session analysis
        biases = {k: v["dominant_bias"] if v else "N/A"
                  for k,v in result.items() if k in ("Asia","London","NewYork")}
        result["cross_analysis"] = {
            "biases": biases,
            "setup": _detect_session_setup(biases),
        }
        return jsonify(result)
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

def _detect_session_setup(biases):
    a = biases.get("Asia","")
    l = biases.get("London","")
    n = biases.get("NewYork","")
    if a == "Accumulation" and l == "Distribution" and n == "Accumulation":
        return "Classic Reversal — Asia acc, London dist, NY breakout ↑"
    if a == "Distribution" and l == "Accumulation" and n == "Distribution":
        return "Classic Reversal — Asia dist, London acc, NY breakout ↓"
    if a in ("Consolidation","Accumulation") and l == "Accumulation" and n == "Accumulation":
        return "Trend Day ↑ — consistent accumulation across sessions"
    if a in ("Consolidation","Distribution") and l == "Distribution" and n == "Distribution":
        return "Trend Day ↓ — consistent distribution across sessions"
    if a == "Consolidation" and l == "Consolidation":
        return "Range Day — waiting for NY session to break direction"
    if a == "Distribution" and l == "Accumulation":
        return "London Reversal — Asia sold, London bought back"
    if a == "Accumulation" and l == "Distribution":
        return "London Fade — Asia built up, London distributed"
    return "Mixed signals — monitor NY open for direction"


# ── EOD P&L Export + Calendar ─────────────────────────────────────────────────
@dashboard_bp.route("/api/eod_export", methods=["POST"])
def api_eod_export():
    e = _auth();
    if e: return e
    try:
        import io, json
        from datetime import datetime, date
        from core.database import get_closed_positions, get_closed_summary, get_recent_trades
        from core.excel_export import export_trades_excel

        today      = date.today().isoformat()
        req_date   = (request.json or {}).get("date", today)

        # Get today's closed positions
        all_closed = get_closed_positions(500)
        day_closed = [p for p in all_closed if (p.get("closed_at","") or "").startswith(req_date)]
        all_trades = get_recent_trades(200)
        day_trades = [t for t in all_trades if (t.get("timestamp","") or "").startswith(req_date)]

        # Daily P&L
        day_pnl   = sum(p.get("pnl",0) or 0 for p in day_closed)
        day_wins  = sum(1 for p in day_closed if (p.get("pnl") or 0) > 0)
        day_loss  = sum(1 for p in day_closed if (p.get("pnl") or 0) < 0)

        # Build EOD Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb  = openpyxl.Workbook()

        # Sheet 1: Daily Summary
        ws1 = wb.active
        ws1.title = f"EOD {req_date}"
        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(color="FFFFFF", bold=True)

        ws1.append(["EOD TRADING JOURNAL", req_date])
        ws1.append([])
        ws1.append(["DAILY SUMMARY"])
        ws1.append(["Date", req_date])
        ws1.append(["Day P&L", f"${day_pnl:.2f}"])
        ws1.append(["Trades Closed", len(day_closed)])
        ws1.append(["Winners", day_wins])
        ws1.append(["Losers",  day_loss])
        ws1.append(["Win Rate", f"{(day_wins/len(day_closed)*100):.1f}%" if day_closed else "0%"])
        ws1.append([])
        ws1.append(["CLOSED POSITIONS"])
        ws1.append(["Symbol","Side","Qty","Entry","Exit","P&L","P&L %","Time"])
        for p in day_closed:
            ws1.append([
                p.get("symbol",""), p.get("side",""),
                p.get("qty",0), p.get("entry_price",0), p.get("exit_price",0),
                p.get("pnl",0), p.get("pnl_pct",0), p.get("closed_at",""),
            ])
        ws1.append([])
        ws1.append(["ALL SIGNALS TODAY"])
        ws1.append(["Time","Symbol","Action","Qty","Status"])
        for t in day_trades:
            ws1.append([t.get("timestamp",""), t.get("symbol",""),
                        t.get("action",""), t.get("quantity",0), t.get("status","")])

        # Sheet 2: Calendar summary (all days)
        ws2 = wb.create_sheet("P&L Calendar")
        ws2.append(["Date","Trades","Winners","Losers","Day P&L","Cumulative P&L"])
        all_p  = get_closed_positions(1000)
        # Group by date
        by_day = {}
        for p in all_p:
            d = (p.get("closed_at","") or "")[:10]
            if d not in by_day: by_day[d] = []
            by_day[d].append(p)
        cum = 0
        for d in sorted(by_day.keys()):
            ps   = by_day[d]
            dpnl = sum(p.get("pnl",0) or 0 for p in ps)
            wins = sum(1 for p in ps if (p.get("pnl") or 0)>0)
            loss = sum(1 for p in ps if (p.get("pnl") or 0)<0)
            cum += dpnl
            ws2.append([d, len(ps), wins, loss, round(dpnl,2), round(cum,2)])

        # Autosize
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len+4, 30)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        from flask import send_file
        return send_file(out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"EOD_Journal_{req_date}.xlsx")
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500


@dashboard_bp.route("/api/pnl_calendar")
def api_pnl_calendar():
    e = _auth();
    if e: return e
    try:
        from core.database import get_closed_positions
        date_from = request.args.get("date_from", "")
        date_to   = request.args.get("date_to",   "")
        all_p  = get_closed_positions(5000)
        # Apply date filter if provided
        if date_from or date_to:
            df = date_from or "2000-01-01"
            dt = date_to   or "2099-12-31"
            all_p = [p for p in all_p if df <= (p.get("closed_at","") or "")[:10] <= dt]
        by_day = {}
        for p in all_p:
            d = (p.get("closed_at","") or "")[:10]
            if not d: continue
            if d not in by_day:
                by_day[d] = {"date":d,"pnl":0,"trades":0,"wins":0,"losses":0}
            by_day[d]["trades"] += 1
            pnl = p.get("pnl") or 0
            by_day[d]["pnl"]    = round(by_day[d]["pnl"] + pnl, 2)
            if pnl > 0: by_day[d]["wins"]   += 1
            else:       by_day[d]["losses"]  += 1
        cal = sorted(by_day.values(), key=lambda x: x["date"])
        cum = 0
        for d in cal:
            cum += d["pnl"]; d["cumulative"] = round(cum,2)
        return jsonify({"calendar": cal, "total_days": len(cal),
                        "cumulative_pnl": round(cum,2),
                        "date_from": date_from, "date_to": date_to})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
